from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl

# Location of chromedriver.exe
chromedriver = 'C:\Program Files\chromedriver.exe'

# Below fixes issues with the dynamic webpage replacing HTML content
options = webdriver.ChromeOptions()
options.add_argument('headless')

browser= webdriver.Chrome(executable_path=chromedriver, chrome_options=options)

# Website to scrape from
url = "https://lotto.bclc.com/winning-numbers.html"

# Grabs HTML from Website in order to parse
browser.get(url)
content = browser.page_source
soup = BeautifulSoup(content, 'html.parser')

# Parsing in order to find the lotto numbers
lotto_nums = soup.find(class_="span29 lotto-649-numbers").find("li").get_text()

lotto_nums = list(lotto_nums.split(' ')) # converts string to list
lotto_nums.pop() # remove the last element which was ' '

# Open existing excel file sheet
wb = openpyxl.load_workbook('lottodata.xlsx')
s = wb['6 49']

# Read text file to know where to insert new data
row_number = 0
f = open("Line_store.txt", "r")
row_number = int(f.readline())
f.close()

# writing to excel sheet
s.cell(row_number, 2, lotto_nums[0])
s.cell(row_number, 3, lotto_nums[1])
s.cell(row_number, 4, lotto_nums[2])
s.cell(row_number, 5, lotto_nums[3])
s.cell(row_number, 6, lotto_nums[4])
s.cell(row_number, 7, lotto_nums[5])

wb.save('lottodata.xlsx') 

# Update text file for next use
f = open("Line_store.txt", "w")
f.write(str(row_number+1))
f.close()

exit()