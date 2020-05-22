import smtplib
import sys
import openpyxl

# Read text file to know where to grab new data
row_number = 0
f = open("Line_store.txt", "r")
row_number = int(f.readline()) - 1
f.close()

# Open existing excel file sheet
wb = openpyxl.load_workbook('lottodata.xlsx', data_only=True)
s = wb['6 49']

# Read numbers from excel
num1 = s.cell(row_number,2).value
num2 = s.cell(row_number,3).value
num3 = s.cell(row_number,4).value
num4 = s.cell(row_number,5).value
num5 = s.cell(row_number,6).value
num6 = s.cell(row_number,7).value

#Sender and receiver
FROM = "patbeqo@gmail.com"
TO = ["patbeqo@gmail.com","pbeqo@uwaterloo.ca"] # must be a list

#Subject for the email
SUBJECT = "Your weekly lotto numbers"

#The email content
TEXT = "Your 6 numbers are: \n\nNum1: {}\nNum2: {}\nNum3: {}\nNum4: {}\nNum5: {}\nNum6: {}".format(num1,num2,num3,num4,num5,num6)

# Formating the email
message = 'Subject: {}\n\n{}'.format(SUBJECT,TEXT)

# Send the mail
server = smtplib.SMTP("smtp.gmail.com", 587)
server.starttls()
server.login("username@gmail.com", "password")
server.sendmail(FROM, TO, message)
server.quit()

exit()